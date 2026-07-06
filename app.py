import os
import io
import streamlit as st
import pandas as pd
from dotenv import load_dotenv
from moodle_client import MoodleClient
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()

MOODLE_URL = os.getenv("MOODLE_URL")
MOODLE_TOKEN = os.getenv("MOODLE_TOKEN")

st.set_page_config(page_title="Seguimiento Moodle", layout="centered")

st.title("Seguimiento de estudiantes en Moodle")

if not MOODLE_URL or not MOODLE_TOKEN:
    st.error("No se encontró la configuración de Moodle. Creá un archivo `.env` con:\n\n"
             "MOODLE_URL=https://tusitio.com/moodle\nMOODLE_TOKEN=tu-token")
    st.stop()

client = MoodleClient(MOODLE_URL, MOODLE_TOKEN)


def generar_excel_con_estilo(df, resumen_df):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Reporte")
        resumen_df.to_excel(writer, index=False, sheet_name="Resumen")

    output.seek(0)
    wb = load_workbook(output)

    fuente = Font(name="Montserrat", size=10)
    fuente_header = Font(name="Montserrat", size=11, bold=True)
    fill_header = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    fill_blanco = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill_gris = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )

    def _get_merge_ranges(values):
        ranges = []
        start = 0
        for i in range(1, len(values)):
            if values[i] != values[start]:
                if i - start > 1:
                    ranges.append((start, i - 1))
                start = i
        if len(values) - start > 1:
            ranges.append((start, len(values) - 1))
        return ranges

    def _style_hoja(ws, df_data, columna_porcentaje=None):
        ncols = ws.max_column
        nrows = ws.max_row

        for col_idx in range(1, ncols + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = fuente_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = thin_border

        for row_idx in range(2, nrows + 1):
            fill = fill_gris if row_idx % 2 == 0 else fill_blanco
            for col_idx in range(1, ncols + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = fuente
                cell.fill = fill
                cell.alignment = align_left if col_idx <= 2 else align_center
                cell.border = thin_border

        for col_idx in (1, 2):
            col_name = df_data.columns[col_idx - 1]
            merge_ranges = _get_merge_ranges(df_data[col_name].tolist())
            for start, end in merge_ranges:
                ws.merge_cells(
                    start_row=start + 2, start_column=col_idx,
                    end_row=end + 2, end_column=col_idx,
                )

        if columna_porcentaje:
            for col_idx in range(1, ncols + 1):
                if ws.cell(row=1, column=col_idx).value == columna_porcentaje:
                    for row_idx in range(2, nrows + 1):
                        ws.cell(row=row_idx, column=col_idx).number_format = "0%"

        for col_idx in range(1, ncols + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
                val = row[0]
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[col_letter].width = min(max_len + 3, 55)

        ws.auto_filter.ref = ws.dimensions

    _style_hoja(
        wb["Reporte"], df,
        columna_porcentaje="Porcentaje Progreso",
    )
    _style_hoja(wb["Resumen"], resumen_df)

    styled_output = io.BytesIO()
    wb.save(styled_output)
    styled_output.seek(0)
    return styled_output


uploaded_file = st.file_uploader(
    "Subí el archivo CSV de estudiantes",
    type="csv",
    help="Formato: nombre;apellido;email;dni (separado por punto y coma)",
)

if uploaded_file is not None:
    df_input = pd.read_csv(uploaded_file, sep=";", header=0, dtype=str, encoding="latin1").fillna("")
    total_filas = len(df_input)
    st.caption(f"Filas en el archivo: {total_filas} (las filas vacías se omiten al procesar)")

    if st.button("Ejecutar seguimiento", type="primary"):
        bar = st.progress(0, text="Iniciando...")
        status = st.empty()

        total = 0
        resultados = []

        for update in client.process_csv(df_input):
            if update["tipo"] == "inicio":
                total = update["total"]
            elif update["tipo"] == "status":
                status.info(update["mensaje"])
                procesados = update.get("procesados", 0)
                bar.progress(
                    min(procesados / max(total, 1), 1.0),
                    text=f"Procesando estudiante {procesados} de {total}",
                )
            elif update["tipo"] == "completo":
                resultados = update["rows"]

        if resultados:
            df_out = pd.DataFrame(resultados)

            total_estudiantes = df_out["Nombre y apellido"].nunique()
            cursos_unicos = [c for c in df_out["Curso"].unique() if c]
            total_cursos = len(cursos_unicos)
            prom_progreso = df_out["Porcentaje Progreso"].mean()

            col1, col2, col3 = st.columns(3)
            col1.metric("Estudiantes procesados", total_estudiantes)
            col2.metric("Cursos detectados", total_cursos)
            col3.metric("Progreso promedio", f"{prom_progreso:.0%}",
                        help="Promedio del porcentaje de avance de todos los estudiantes en todos los cursos")

            mask_con_curso = df_out["Curso"] != ""
            if mask_con_curso.any():
                resumen_curso = (
                    df_out[mask_con_curso]
                    .groupby("Curso", sort=False)
                    .agg(
                        Inscriptos=("Nombre y apellido", "nunique"),
                        **{"Sin inicio": ("Porcentaje Progreso", lambda x: (x == 0).sum())},
                        **{"% En curso": ("Porcentaje Progreso", lambda x: ((x > 0) & (x < 1.0)).sum())},
                        Completos=("Porcentaje Progreso", lambda x: (x >= 1.0).sum()),
                    )
                    .reset_index()
                )

                st.subheader("Resumen por curso")
                st.dataframe(resumen_curso, use_container_width=True, hide_index=True)

                st.success(f"Reporte generado: {len(df_out)} registros")
                st.dataframe(df_out, use_container_width=True, hide_index=True)

                excel_data = generar_excel_con_estilo(df_out, resumen_curso)

                st.download_button(
                    label="Descargar Excel",
                    data=excel_data,
                    file_name="reporte_seguimiento.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            else:
                st.warning("No se encontraron cursos para generar el reporte.")
        else:
            st.warning("No se generaron resultados.")
else:
    st.info("Subí un archivo CSV para comenzar.")
